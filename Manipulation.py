from sqlite3 import Row
from openpyxl import Workbook, load_workbook
import datetime
import os

import openpyxl

path_file = "sample.xlsx"

#wb = Workbook(); # filesystem
wb = load_workbook('sample.xlsx') # load file
ws = wb.active # active worksheet
sh = wb['Sheet']

ws['A2'] = 1
ws.append([1,2,3]) # creating tree columns
ws['A2'] = datetime.datetime.now()

for row in ws.iter_rows(min_row=1, min_col=1, max_row=12, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print

ramal_finished = ''
status_ramal = "Search status of ramal..."

for cell in sh['A']:
    if sh['A'] == ramal_finished :
        print("Não existe ramal para realização da busca!")
    else:
        sh['B'] = status_ramal
        print('All be right!')
        
# INPUT VARIABLE HERE
sh.cell(2,2).value = status_ramal

print(sh.max_row) # use for count how much ramais exist
print(sh.max_column) # used for know how columns exist

## READ ROW + COLUM
for i in range(1, sh.max_row+1):
    for cell in sh[i]:
        print(cell.value)

for row in sh.iter_rows(1, sh.max_row+1):
    for cell in row:
        print(cell.value)

wb.save('sample.xlsx')
wb.close